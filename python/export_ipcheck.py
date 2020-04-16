# -*- coding:utf8 -*
import redis
from openpyxl import Workbook
import sys
import os
import time
import datetime
import platform
import rank_user
import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer,encoding='utf-8')

class Logger:
	def info(self, *args):
		print('[{}][info] {}'.format(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),*args))
	def error(self, *args):
		print('[{}][error] {}'.format(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),*args))
	def warn(self, *args):
		print('[{}][warn] {}'.format(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),*args))
	def debug(self, *args):
		print('[{}][debug] {}'.format(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),*args))
		

logger = Logger()

host = '192.168.82.164'
r = redis.Redis(host=host, port=6379)
title = ['省份','市','用户数量']
years = ['2017', '2018', '2019']
wb = Workbook()
ws = wb.active
ws.title = "付费玩家报告期内按年IP地址分布分析"
#chr : 数字转 ascii
#ord : ascii 转数字

idx = 65
for year in years:
	logger.info("deal with {}".format(year))
	for tx in title:
		cell_idx = chr(idx) + str('1')
		ws.row_dimensions[1].height = 30
		ws.column_dimensions[chr(idx)].width = 20
		ws[cell_idx] = tx
		idx = idx + 1

	rkey = "ip_counter:" + year
	
	res = r.zrevrange(rkey, 0, -1, 'withscores')
	row = 2
	for k, v in res:
		s = k.decode('utf-8')
		arr = s.split(':', 1)
		prov = arr[0]
		city = arr[1]
		tmp_idx = idx - len(title)
		print('res', prov, city, v, year)		
		cell_idx = chr(tmp_idx) + str(row)
		logger.debug('cell_idx:{}'.format(cell_idx))
		ws[cell_idx] = prov 
		tmp_idx = tmp_idx + 1
		cell_idx = chr(tmp_idx) + str(row)
		ws[cell_idx] = city 
		tmp_idx = tmp_idx + 1
		cell_idx = chr(tmp_idx) + str(row)
		ws[cell_idx] = v
		tmp_idx = tmp_idx + 1
		row = row + 1
	
	idx = idx + 2

wb.save("付费玩家报告期内按年IP地址分布分析.xlsx")
logger.debug("all done !!")



