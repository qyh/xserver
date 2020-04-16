# -*- coding:utf8 -*
import redis
from openpyxl import Workbook
import sys
import os
import time
import datetime
import platform
import rank_user
import sys
import io
import MySQLdb
#sys.stdout = io.TextIOWrapper(sys.stdout.buffer,encoding='utf-8')
os.environ['NLS_LANG'] = 'SIMPLIFIED CHINESE_CHINA.UTF8'
class Logger:
	def info(self, *args):
		print('[{}][info] {}'.format(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),*args))
	def error(self, *args):
		print('[{}][error] {}'.format(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),*args))
	def warn(self, *args):
		print('[{}][warn] {}'.format(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),*args))
	def debug(self, *args):
		print('[{}][debug] {}'.format(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),*args))
		

logger = Logger()

host = '192.168.82.164'
r = redis.Redis(host=host, port=6379)
title = ['日期','游戏名称','用户ID', '昵称', '游戏局数']
wb = Workbook()
ws = wb.active
ws.title = "每天牌局top100玩家2018"
#chr : 数字转 ascii
#ord : ascii 转数字

# 打开数据库连接
db = MySQLdb.connect(host="1475177020277729-73404573.cn-shenzhen.datalakeanalytics.aliyuncs.com",port=10000, user="qinyanhong_s1475177020277729",passwd="wQOsikfKJLctvn1ctc",db="oss_gamelog2017" )

# 使用cursor()方法获取操作游标 
cursor = db.cursor()

idx = 65
for tx in title:
	cell_idx = chr(idx) + str('1')
	#ws.row_dimensions[1].height = 30
	#ws.column_dimensions[chr(idx)].width = 20
	ws[cell_idx] = tx
	idx = idx + 1
row = 2


f = open('game_record_rank_2018.txt', 'rb')
l = f.readlines()
for i in range(0, len(l)):
	lineStr = l[i].decode("utf-8")
	line = lineStr[0:-1].split(",")
	idx = 65
	cell_idx = chr(idx) + str(row)
	ws[cell_idx] = line[0]
	idx = idx + 1
	cell_idx = chr(idx) + str(row)
	ws[cell_idx] = '桂林字牌'
	idx = idx + 1
	cell_idx = chr(idx) + str(row)
	ws[cell_idx] = line[1]
	idx = idx + 1
	cell_idx = chr(idx) + str(row)
	ws[cell_idx] = line[2]
	idx = idx + 1
	cell_idx = chr(idx) + str(row)
	ws[cell_idx] = line[3]
	row = row + 1
wb.save(ws.title + ".xlsx")
logger.debug("all done !!")



