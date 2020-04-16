# coding: utf-8
import redis
from openpyxl import Workbook
import sys
import os
import time
import datetime
import platform
import rank_user
import sys
import io
import MySQLdb

title = ['日期','游戏名称','用户ID', '昵称', '游戏局数']
wb = Workbook()
ws = wb.active
ws.title = "每天牌局top100玩家2017"
#chr : 数字转 ascii
#ord : ascii 转数字

# 打开数据库连接
db = MySQLdb.connect(host="1475177020277729-73404573.cn-shenzhen.datalakeanalytics.aliyuncs.com",port=10000, user="qinyanhong_s1475177020277729",passwd="wQOsikfKJLctvn1ctc",db="oss_gamelog2017",charset='utf8')

# 使用cursor()方法获取操作游标 
cursor = db.cursor()

def getNextDay(str):
	_t = datetime.datetime.strptime(str, "%Y-%m-%d")
	out_date = (_t + datetime.timedelta(days=1)).strftime("%Y-%m-%d")
	return out_date

idx = 65
for tx in title:
	cell_idx = chr(idx) + str('1')
	ws[cell_idx] = tx
	idx = idx + 1
row = 2
start_time_str = "2017-03-05"
end_time_str = "2018-01-01"

e_time = time.mktime(time.strptime(end_time_str,"%Y-%m-%d"))
start_t = datetime.datetime.strptime(start_time_str, "%Y-%m-%d")
out_date = (start_t + datetime.timedelta(days=7)).strftime("%Y-%m-%d")
s_time = time.mktime(time.strptime(start_time_str,"%Y-%m-%d"))
tmp_time = s_time
tmp_str = start_time_str
year = 2017
topList = []
while tmp_time <= e_time:
	print("tmp_str:",tmp_str)
	_str = tmp_str.replace("-", "_")
	# 每张表查询牌局 top 10
	sql = 'SELECT date_format(starttime, "%Y-%m-%d") as d, userid, nickname, count(*) as s FROM (select userid,startTime,nickname from oss_gamelog{}.gameuserinfolog_{} where starttime < \'{}\' union select userid,startTime,nickname from oss_gamelog{}.matchuserinfolog_{} where starttime <\'{}\' union select userid,startTime,nickname from oss_gamelog{}.newroomselfuserinfolog_{} where starttime <\'{}\') group by date_format(starttime, "%Y-%m-%d"), userid, nickname order by s desc  limit 10;'.format(year,_str, end_time_str,year,_str, end_time_str,year,_str,end_time_str)
	print("sql:", sql)
	cursor.execute(sql)
	# 获取记录并放入topList
	results = cursor.fetchall()
	for line in results:
		item = {}
		#用户ID
		item['userid'] = line[1]
		#日期
		item['time'] = line[0]
		#牌局数
		item['num'] = line[3]
		#昵称
		item['nickname'] = line[2]
		#对应表后缀
		item['prefix'] = _str
		#对应年份
		item['year'] = year
		#放入topList
		topList.append(item)
		#写入excel
		idx = 65
		cell_idx = chr(idx) + str(row)
		ws[cell_idx] = item['time']
		idx = idx + 1
		
		cell_idx = chr(idx) + str(row)
		ws[cell_idx] = '桂林字牌'
		idx = idx + 1
		
		cell_idx = chr(idx) + str(row)
		ws[cell_idx] = item['userid']
		idx = idx + 1
		
		cell_idx = chr(idx) + str(row)
		ws[cell_idx] = item['nickname']
		idx = idx + 1
		
		cell_idx = chr(idx) + str(row)
		ws[cell_idx] = item['num']
		idx = idx + 1
		
		row = row + 1
	tmp_t = datetime.datetime.strptime(tmp_str, "%Y-%m-%d")
	tmp_str = (tmp_t + datetime.timedelta(days=7)).strftime("%Y-%m-%d")
	tmp_time = time.mktime(time.strptime(tmp_str,"%Y-%m-%d"))

wb.save(ws.title + ".xlsx")

#按牌局数倒序排序
topList.sort(key=lambda item: item['num'],reverse=True)
n=0
#查前100明细
f = open(ws.title + "_明细.txt", "wb")
f.write("用户ID,用户名,昵称,变化前金币,变化后金币,服务费,开始时间,结束时间,等级,倍率,状态,类型(0倍率,1比赛,2私人场)\n".encode('utf-8'))
for item in topList:
	if n >= 100:
		break
	print(item)
	#查询用户倍率明细
	sql = "select userid, username, nickname, beforegoldcoin, aftergoldcoin, servicecharge, starttime, endtime, `level`, roomrate, state, 0 as `type` from oss_gamelog{}.gameuserinfolog_{} where userid={} and starttime>='{}' and starttime <'{}' ".format(item['year'], item['prefix'], item['userid'],item['time'], getNextDay(item['time']))
	print('gameuserinfolog_ detail sql:', sql)
	cursor.execute(sql)
	res = cursor.fetchall()
	for line in res:
		txt = "{},{},{},{},{},{},{},{},{},{},{},{}\n".format(line[0],line[1],line[2],line[3],line[4],line[5],line[6],line[7],line[8],line[9],line[10], line[11])
		f.write(txt.encode('utf-8'))
	#查询用户比赛明细
	sql = "select userid, username, nickname, beforeintegral, afterintegral, 0 as servicecharge, starttime, endtime, `level`, 0 as deskrate, state, 1 as `type` from oss_gamelog{}.matchuserinfolog_{} where userid={} and starttime>='{}' and starttime <'{}' ".format(item['year'], item['prefix'], item['userid'],item['time'], getNextDay(item['time']))
	print('matchuserinfolog_ detail sql:', sql)
	cursor.execute(sql)
	res = cursor.fetchall()
	for line in res:
		txt = "{},{},{},{},{},{},{},{},{},{},{},{}\n".format(line[0],line[1],line[2],line[3],line[4],line[5],line[6],line[7],line[8],line[9],line[10], line[11])
		f.write(txt.encode('utf-8'))

	#查询用户私人场明细
	sql = "select userid, username, nickname, beforegoldcoin, aftergoldcoin, servicecharge, starttime, endtime,`level`, deskrate, state, 2 as `type` from oss_gamelog{}.newroomselfuserinfolog_{} where userid={} and starttime>='{}' and starttime <'{}' ".format(item['year'], item['prefix'], item['userid'],item['time'], getNextDay(item['time']))
	print('newroomselfuserinfolog_ detail sql:', sql)
	cursor.execute(sql)
	res = cursor.fetchall()
	for line in res:
		txt = "{},{},{},{},{},{},{},{},{},{},{},{}\n".format(line[0],line[1],line[2],line[3],line[4],line[5],line[6],line[7],line[8],line[9],line[10], line[11])
		f.write(txt.encode('utf-8'))

	n = n + 1
	print("第"+str(n)+"个玩家处理完毕!")
f.close()
print("all done !!")



