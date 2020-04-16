# -*- coding:utf8 -*
import redis
from openpyxl import Workbook
import sys
import os
import time
import datetime
import platform
import rank_user

class Logger:
	def info(self, *args):
		print('[{}][info] {}'.format(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),*args))
	def error(self, *args):
		print('[{}][error] {}'.format(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),*args))
	def warn(self, *args):
		print('[{}][warn] {}'.format(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),*args))
	def debug(self, *args):
		print('[{}][debug] {}'.format(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),*args))
		

logger = Logger()

host = '192.168.82.164'
r = redis.Redis(host=host, port=6379)

title = ['序号','游戏账号ID','注册时间','注册IP','截止报告期未实时最后登陆时间（汇总数据时点）','截止报告期等级','首次充值时间','最后充值日期','截止报告期充值次数','截止报告期充值累计金额','截止报告期游戏币留存数','截止报告期累计活跃天数','游戏胜局数','游戏负局数','充值金币数','充值对战卡数', '主要登陆IP', '主要充值IP', '主要登录IMEI/MAC', '主要充值IMEI/MAC']
ks = ['id','code', 'regTime','regIp', 'lastLoginTime', 'level', 'firstRechargeTime', 'lastRechargeTime','rechargeCount','totalAmount','goldCoin','activeDay','winCount','loseCount','goldCoinRecharge','gameCardRecharge', 'mainLoginIp', 'mainRechargeIp','mainLoginImei', 'mainRechargeImei']


wb = Workbook()
ws = wb.active
ws.title = "前10000大R基本信息"
#chr : 数字转 ascii
#ord : ascii 转数字
idx = 65
for tx in title:
	cell_idx = chr(idx) + str('1')
	ws.row_dimensions[1].height = 30
	ws.column_dimensions[chr(idx)].width = 20
	ws[cell_idx] = tx
	idx = idx + 1

users = rank_user.get_rank_user()
row = 2
i = 1
for userID, amount in users.items():
	rdsKey = "audit_user:" + userID
	user_info = r.hgetall(rdsKey)
	idx = 65

	#user_info['regIp'.encode('utf8')] = ' '
	user_info['mainLoginIp'.encode('utf8')] = ' '
	user_info['mainLoginImei'.encode('utf8')] = ' '
	user_info['mainRechargeImei'.encode('utf8')] = ' '
	user_info['mainRechargeIp'.encode('utf8')] = ' '
	user_info['id'.encode('utf8')] = i
	i = i + 1
	for _k in ks:
		cell_idx = chr(idx) + str(row)
		ws[cell_idx] = user_info.get(_k.encode('utf8')) or 0
		idx = idx + 1
	row = row + 1

wb.save("前10000大R基本信息.xlsx")
logger.debug("all done !!")



